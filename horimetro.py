import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime
from collections import defaultdict
from pathlib import Path
import os

while True:
    # Puxa os dados do relatorio GRM
    def puxar_dados(nome, caminho):
        if nome == "grm":
            wb = openpyxl.load_workbook(caminho)
            ws = wb.active

            dados = []
            for row in ws.iter_rows(min_row=6, values_only=True):
                tag = row[0]
                data = row[1]
                valor = row[2]
                dados.append([tag, data, valor, ""])  # A observação estará vazia por enquanto
            return dados

        elif nome == "horimetro":
            wb = openpyxl.load_workbook(caminho)
            ws = wb.active

            dados = []
            for row in ws.iter_rows(min_row=3, values_only=True):
                tag = row[0]
                data = row[1]
                valor = row[2]
                if row[3] is None:
                    obs = ""
                else:
                    obs = row[3]

                dados.append([tag, data, valor, obs])
            return dados

        else:
            raise ValueError("Nome inválido. Use 'grm' ou 'horimetro'.")

    def transformar_dados_em_dicionario(dados):
        dados_horimetros_anterior = []
        for linha in dados:
            dicionario = {
                "tag": linha[0],
                "horimetro": linha[2]
            }
            dados_horimetros_anterior.append(dicionario)
        return dados_horimetros_anterior

    # Processar valores e observações
    def processar_dados(dados):
        header = ["Tag", "Data", "Horimetro", "Observação"]
        linhas = dados
        dicionario = transformar_dados_em_dicionario(dados_ultimo_horimetro)

        for coluna in linhas:
            coluna[2] = str(coluna[2])

            if coluna[0].startswith("C"):
                coluna[2] = coluna[2].replace(".", "")

            elif coluna[0].startswith("E") and len(coluna[2]) <= 5:
                coluna[2] = coluna[2].replace(".", "")

            elif coluna[0].startswith("T") and len(coluna[2]) <= 5:
                coluna[2] = coluna[2].replace(".", "")

            if len(coluna[2]) < 4 or len(coluna[2]) > 7:
                coluna[3] = "Fora do padrão"

            for item in dicionario:
                if item["tag"] == coluna[0]:
                    horimetro_antigo = float(item["horimetro"])
                    coluna[2] = float(coluna[2])

                    if horimetro_antigo == coluna[2] and coluna[1][:10].replace("/", "-") == hoje:
                        coluna[3] = "Horimetro igual ao anterior"
                 
                    if coluna[2] < horimetro_antigo:
                        coluna[3] = "Horimetro menor que anterior"
        
        # Retorna os dados processados sem exclusão de duplicados
        dados_processados = [header] + linhas
        return dados_processados

    def data_emissão():
        hoje = datetime.now().strftime('%d-%m-%Y')
        ws.merge_cells("A1:D1")
        ws["A1"] = f"Emitido em {hoje}"
        ws["A1"].font = Font(bold=True)
        ws["A1"].alignment = Alignment(horizontal="center")

    def configuracao_estilos():
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        center_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style="thin"), 
            right=Side(style="thin"), 
            top=Side(style="thin"), 
            bottom=Side(style="thin")
        )

        for i, row in enumerate(dados_planilha_atual, start=2):
            for j, value in enumerate(row, start=1):
                cell = ws.cell(row=i, column=j, value=value)
                if i == 2: 
                    cell.font = header_font
                    cell.fill = header_fill
                cell.alignment = center_alignment
                cell.border = thin_border

    # Ajustar largura das colunas automaticamente
        for col_idx, col in enumerate(ws.iter_cols(min_row=2, max_row=ws.max_row, max_col=ws.max_column), start=1):
            max_length = max(len(str(cell.value or "")) for cell in col)
            col_letter = get_column_letter(col_idx)  # Captura a letra da coluna pelo índice
            ws.column_dimensions[col_letter].width = max_length + 2

    wb = Workbook()
    ws = wb.active
    ws.title = "Horimetros"

    hoje = datetime.now().strftime('%d-%m-%Y')
    downloads_path = Path.home() / "Downloads"
    downloads_path = str(downloads_path)

    caminho = downloads_path + f"\\Horimetros {hoje}.xlsx"

    data_emissão()
    diretorio_relatorio = downloads_path + "//" + input("Informe o nome do arquivo de relatório: ") + ".xlsx"
    diretorio_horimetro = downloads_path + "//" + input("Informe o nome de arquivo do último horimetro: ") + ".xlsx"
    os.system("cls")

    dados_grm = puxar_dados("grm", diretorio_relatorio)
    dados_ultimo_horimetro = puxar_dados("horimetro", diretorio_horimetro)
    transformar_dados_em_dicionario(dados_ultimo_horimetro)
    dados_planilha_atual = processar_dados(puxar_dados("grm", diretorio_relatorio))
    configuracao_estilos()
    wb.save(caminho)
    print(f"Dados salvos com sucesso em {caminho}")

    # Salvar o arquivo
